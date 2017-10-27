from PyQt4 import QtCore, QtGui
from PyQt4 import QtWebKit
import logging
import datetime
from openpyxl.workbook import Workbook
import os
import time
import cx_Oracle
import csv
import pyodbc
import getpass
from time import sleep
import sys
global origQuery



logging.basicConfig(level=logging.DEBUG, filename='tdmt.log', filemode='a')
try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

try:
    _encoding = QtGui.QApplication.UnicodeUTF8
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig, _encoding)
except AttributeError:
    def _translate(context, text, disambig):
        return QtGui.QApplication.translate(context, text, disambig)
searchKey= ''
try:
    
    URL= open('config\\setDirectory.ini').read()
    ENV= open('config\\setEnv.ini').readlines()
    for i in range(len(ENV)):
        ENV[i]=ENV[i].strip()        
    ENV.sort()
    
    MOD= open('config\\setModule.ini').readlines()
    for k in range(len(MOD)):
        MOD[k]=MOD[k].strip()       
    MOD.sort()    
    
except:
    logging.exception(str(datetime.datetime.now())+"[FILE ERROR]:")

adminppwd ="bsctdm2017"
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName(_fromUtf8("MainWindow"))
        MainWindow.resize(871, 616)
        MainWindow.activateWindow()
        MainWindow.setWindowIcon(QtGui.QIcon('config\\tdmt.icon'))
        self.centralwidget = QtGui.QWidget(MainWindow)
        self.centralwidget.setObjectName(_fromUtf8("centralwidget"))
        self.label = QtGui.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(20, 10, 281, 41))
        
        self.label.setText("Test Data Mining Tool")
        self.label.setStyleSheet("font-family:Century Gothic")
        
        font = QtGui.QFont()
        font.setPointSize(17)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName(_fromUtf8("label"))
        self.stackedWidget = QtGui.QStackedWidget(self.centralwidget)
        self.stackedWidget.setGeometry(QtCore.QRect(10, 80, 841, 481))
        self.stackedWidget.setFrameShape(QtGui.QFrame.NoFrame)
        self.stackedWidget.setFrameShadow(QtGui.QFrame.Plain)
        self.stackedWidget.setObjectName(_fromUtf8("stackedWidget"))
        self.loginpg = QtGui.QWidget()
        self.loginpg.setObjectName(_fromUtf8("loginpg"))
        self.login_group = QtGui.QGroupBox(self.loginpg)
        self.login_group.setGeometry(QtCore.QRect(210, 175, 451, 201))
        #self.login_group.setStyleSheet("border:1px solid; border-radius:4px;")
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        font.setPointSize(10)
        self.login_group.setFont(font)
        self.login_group.setAlignment(QtCore.Qt.AlignCenter)
        self.login_group.setObjectName(_fromUtf8("login_group"))
        self.picture = QtGui.QLabel(self.loginpg)
        self.picture.setObjectName(_fromUtf8("picture"))
        self.picture.setGeometry(QtCore.QRect(190, 10, 121, 121))
        self.picture.setPixmap(QtGui.QPixmap('config\\tdmt.png').scaled(120,120,transformMode=QtCore.Qt.SmoothTransformation))
        self.frontBan = QtGui.QLabel(self.loginpg)
        self.frontBan.setGeometry(QtCore.QRect(320, 20, 331, 71))
        self.frontBan.setObjectName(_fromUtf8("frontBan"))
        self.frontBan.setText("Test Data Mining Tool")
        self.frontBan.setStyleSheet("font-family:Century Gothic;font-weight:bold; font-size:24pt")
        
        self.user_button = QtGui.QPushButton(self.login_group)
        self.user_button.setGeometry(QtCore.QRect(76, 80, 121, 41))
        self.user_button.setObjectName(_fromUtf8("user_button"))
        self.admin_button = QtGui.QPushButton(self.login_group)
        self.admin_button.setGeometry(QtCore.QRect(263, 80, 121, 41))
        self.admin_button.setObjectName(_fromUtf8("admin_button"))
        self.adminpwd = QtGui.QLineEdit(self.login_group)
        self.adminpwd.setGeometry(QtCore.QRect(252, 90, 121, 20))
        self.adminpwd.setEchoMode(QtGui.QLineEdit.Password)
        self.adminpwd.setCursorMoveStyle(QtCore.Qt.VisualMoveStyle)
        self.adminpwd.setObjectName(_fromUtf8("adminpwd"))
        self.pwdlabel = QtGui.QLabel(self.login_group)
        self.pwdlabel.setGeometry(QtCore.QRect(250, 60, 63, 16))
        self.pwdlabel.setObjectName(_fromUtf8("pwdlabel"))
        self.go_button = QtGui.QPushButton(self.login_group)
        self.go_button.setGeometry(QtCore.QRect(384, 89, 51, 21))
        self.go_button.setObjectName(_fromUtf8("go_button"))
        self.label_2 = QtGui.QLabel(self.login_group)
        self.label_2.setGeometry(QtCore.QRect(150, 140, 180, 35))
        self.label_2.setObjectName(_fromUtf8("label_2"))
        self.stackedWidget.addWidget(self.loginpg)
        self.searchpg = QtGui.QWidget()
        self.searchpg.setObjectName(_fromUtf8("searchpg"))
        self.clear_button = QtGui.QPushButton(self.searchpg)
        self.clear_button.setGeometry(QtCore.QRect(740, 20, 75, 21))
        self.clear_button.setAutoDefault(False)
        self.clear_button.setObjectName(_fromUtf8("clear_button"))        
        
        self.search_button = QtGui.QPushButton(self.searchpg)
        self.search_button.setGeometry(QtCore.QRect(546, 19, 75, 23))
        self.adminpwd.returnPressed.connect(self.go_button.click)
        self.search_button.setAutoDefault(True)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.search_button.setFont(font)
        self.search_button.setCheckable(False)
        self.search_button.setAutoRepeat(False)
        self.search_button.setFlat(False)
        self.search_button.setObjectName(_fromUtf8("search_button"))
        
        self.add_button = QtGui.QPushButton(self.searchpg)
        self.add_button.setGeometry(QtCore.QRect(641, 19, 75, 23))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.add_button.setFont(font)
        self.add_button.setCheckable(False)
        self.add_button.setChecked(False)
        self.add_button.setAutoDefault(True)
        self.add_button.setObjectName(_fromUtf8("add_button"))
        self.searchkey_lineedit = QtGui.QLineEdit(self.searchpg)
        self.searchkey_lineedit.setGeometry(QtCore.QRect(57, 20, 268, 20))
        self.searchkey_lineedit.setFrame(True)
        self.searchkey_lineedit.setObjectName(_fromUtf8("searchkey_lineedit"))
        self.searchkey_lineedit.returnPressed.connect(self.search_button.click)

        self.module_combobox2 = QtGui.QComboBox(self.searchpg)
        self.module_combobox2.setGeometry(QtCore.QRect(380, 20, 151, 20))
        self.module_combobox2.setFrame(True)
        self.module_combobox2.setObjectName(_fromUtf8("module_combobox"))
        self.module_combobox2.setStyleSheet("padding-left:5px;")
        
        self.resultspane_listwidget = QtGui.QListWidget(self.searchpg)
        self.resultspane_listwidget.setGeometry(QtCore.QRect(20, 66, 801, 371))
        self.resultspane_listwidget.setStyleSheet(_fromUtf8("font: 10pt \"MS Shell Dlg 2\";\n"
""))
        self.resultspane_listwidget.setFrameShape(QtGui.QFrame.NoFrame)
        self.resultspane_listwidget.setFrameShadow(QtGui.QFrame.Sunken)
        self.resultspane_listwidget.setObjectName(_fromUtf8("resultspane_listwidget"))
        self.resultspane_listwidget.setSelectionMode(QtGui.QAbstractItemView.ExtendedSelection)
        self.resultspane_listwidget.setSortingEnabled(True)
        self.resultspane_label = QtGui.QLabel(self.searchpg)
        self.resultspane_label.setGeometry(QtCore.QRect(20, 50, 46, 13))
        self.resultspane_label.setObjectName(_fromUtf8("resultspane_label"))
        self.delete_button = QtGui.QPushButton(self.searchpg)
        self.delete_button.setGeometry(QtCore.QRect(750, 450, 75, 23))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.delete_button.setFont(font)
        self.delete_button.setCheckable(False)
        self.delete_button.setChecked(False)
        self.delete_button.setAutoDefault(True)
        self.delete_button.setObjectName(_fromUtf8("delete_button"))
        self.searchhere_label = QtGui.QLabel(self.searchpg)
        self.searchhere_label.setGeometry(QtCore.QRect(20, 21, 81, 16))
        self.moduletext_label = QtGui.QLabel(self.searchpg)
        self.moduletext_label.setGeometry(QtCore.QRect(330, 20, 61, 20))
        self.moduletext_label.setText(" Module: ")
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.searchhere_label.setFont(font)
        self.searchhere_label.setObjectName(_fromUtf8("searchhere_label"))
        self.runstatus01 = QtGui.QLabel(self.searchpg)
        self.runstatus01.setGeometry(QtCore.QRect(20, 450, 261, 16))
        self.runstatus01.setObjectName(_fromUtf8("runstatus01"))
        self.stackedWidget.addWidget(self.searchpg)
        self.searchentitypg = QtGui.QWidget()
        self.searchentitypg.setObjectName(_fromUtf8("searchentitypg"))
        self.query_label = QtGui.QLabel(self.searchentitypg)
        self.query_label.setGeometry(QtCore.QRect(12, 92, 71, 31))
        self.query_label.setObjectName(_fromUtf8("query_label"))
        self.module_label = QtGui.QLabel(self.searchentitypg)
        self.module_label.setGeometry(QtCore.QRect(12, 30, 51, 20))
        self.module_label.setObjectName(_fromUtf8("module_label"))
        self.desc_lineedit = QtGui.QLineEdit(self.searchentitypg)
        self.desc_lineedit.setGeometry(QtCore.QRect(80, 70, 501, 20))
        self.desc_lineedit.setFrame(True)
        self.desc_lineedit.setObjectName(_fromUtf8("desc_lineedit"))

        self.permi_label = QtGui.QLabel(self.searchentitypg)
        self.permi_label.setGeometry(QtCore.QRect(340, 30, 51, 20))
        self.permi_label.setObjectName(_fromUtf8("permi_label"))
        self.permi_label.setText("Permissions: ")
        
        self.permi_lineedit = QtGui.QLineEdit(self.searchentitypg) #permi
        self.permi_lineedit.setGeometry(QtCore.QRect(400, 31, 180, 20))
        self.permi_lineedit.setFrame(True)
        self.permi_lineedit.setObjectName(_fromUtf8("permi_lineedit"))
        self.permi_lineedit.setText("All")
        self.permi_lineedit.setToolTip("Enter user ids seperated with commas E.g: jwdaoe01, sdivaa02")
        
        self.desc_label = QtGui.QLabel(self.searchentitypg)
        self.desc_label.setGeometry(QtCore.QRect(10, 70, 81, 20))
        self.desc_label.setObjectName(_fromUtf8("desc_label"))
        self.dbcreds_group = QtGui.QGroupBox(self.searchentitypg)
        self.dbcreds_group.setGeometry(QtCore.QRect(599, 3, 221, 111))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.dbcreds_group.setFont(font)
        self.dbcreds_group.setObjectName(_fromUtf8("dbcreds_group"))
        self.userId_lineedit = QtGui.QLineEdit(self.dbcreds_group)
        self.userId_lineedit.setGeometry(QtCore.QRect(89, 20, 111, 20))
        self.userId_lineedit.setFrame(True)
        self.userId_lineedit.setObjectName(_fromUtf8("userId_lineedit"))
        self.pwd_label = QtGui.QLabel(self.dbcreds_group)
        self.pwd_label.setGeometry(QtCore.QRect(20, 50, 71, 20))
        self.pwd_label.setObjectName(_fromUtf8("pwd_label"))
        self.envdrop_combobox = QtGui.QComboBox(self.dbcreds_group)
        self.envdrop_combobox.setGeometry(QtCore.QRect(89, 78, 111, 22))
        self.envdrop_combobox.setFrame(False)
        self.envdrop_combobox.setObjectName(_fromUtf8("envdrop_combobox"))
        self.envdrop_combobox.setStyleSheet("padding-left:5px;")
        self.usrname_label = QtGui.QLabel(self.dbcreds_group)
        self.usrname_label.setGeometry(QtCore.QRect(20, 20, 71, 16))
        self.usrname_label.setObjectName(_fromUtf8("usrname_label"))
        self.env_label = QtGui.QLabel(self.dbcreds_group)
        self.env_label.setGeometry(QtCore.QRect(20, 80, 70, 16))
        self.env_label.setObjectName(_fromUtf8("env_label"))
        self.password_lineedit = QtGui.QLineEdit(self.dbcreds_group)
        self.password_lineedit.setGeometry(QtCore.QRect(89, 50, 111, 20))
        self.password_lineedit.setInputMask(_fromUtf8(""))
        self.password_lineedit.setFrame(True)
        self.password_lineedit.setEchoMode(QtGui.QLineEdit.Password)
        self.password_lineedit.setObjectName(_fromUtf8("password_lineedit"))
        self.save_button = QtGui.QPushButton(self.searchentitypg)
        self.save_button.setGeometry(QtCore.QRect(10, 444, 75, 23))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.save_button.setFont(font)
        self.save_button.setObjectName(_fromUtf8("save_button"))
        self.copytoclipboard_button = QtGui.QPushButton(self.searchentitypg)
        self.copytoclipboard_button.setGeometry(QtCore.QRect(110, 444, 101, 23))
        self.copytoclipboard_button.setObjectName(_fromUtf8("copytoclipboard_button"))
        self.update_button = QtGui.QPushButton(self.searchentitypg)
        self.update_button.setGeometry(QtCore.QRect(240, 444, 75, 23))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.update_button.setFont(font)
        self.update_button.setObjectName(_fromUtf8("update_button"))
        self.execute_button = QtGui.QPushButton(self.searchentitypg)
        self.execute_button.setGeometry(QtCore.QRect(718, 444, 101, 23))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.execute_button.setFont(font)
        self.execute_button.setObjectName(_fromUtf8("execute_button"))
        self.module_combobox = QtGui.QComboBox(self.searchentitypg)
        self.module_combobox.setGeometry(QtCore.QRect(80, 31, 501, 20))
        self.module_combobox.setFrame(False)
        self.module_combobox.setObjectName(_fromUtf8("module_combobox"))
        self.module_combobox.setStyleSheet("padding-left:5px;")
        self.rowcount_label = QtGui.QLabel(self.searchentitypg)
        self.rowcount_label.setGeometry(QtCore.QRect(566, 446, 61, 20))
        self.rowcount_label.setObjectName(_fromUtf8("rowcount_label"))
        self.rowcount_lineedit = QtGui.QLineEdit(self.searchentitypg)
        self.rowcount_lineedit.setGeometry(QtCore.QRect(633, 446, 61, 20))
        self.rowcount_lineedit.setFrame(True)
        self.rowcount_lineedit.setObjectName(_fromUtf8("rowcount_lineedit"))
        self.querydata_label = QtGui.QLabel(self.searchentitypg)
        self.querydata_label.setGeometry(QtCore.QRect(12, 2, 81, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.querydata_label.setFont(font)
        self.querydata_label.setObjectName(_fromUtf8("querydata_label"))
        self.clearinputs_button = QtGui.QPushButton(self.searchentitypg)
        self.clearinputs_button.setGeometry(QtCore.QRect(340, 444, 75, 21))
        self.clearinputs_button.setAutoDefault(False)
        self.clearinputs_button.setObjectName(_fromUtf8("clearinputs_button"))
        self.query_plaintextedit = QtGui.QPlainTextEdit(self.searchentitypg)
        self.query_plaintextedit.setGeometry(QtCore.QRect(10, 122, 811, 311))
        self.query_plaintextedit.setFrameShape(QtGui.QFrame.NoFrame)
        self.query_plaintextedit.setFrameShadow(QtGui.QFrame.Plain)
        self.query_plaintextedit.setObjectName(_fromUtf8("query_plaintextedit"))
        self.dbcreds_group.raise_()
        self.query_label.raise_()
        self.module_label.raise_()
        self.desc_lineedit.raise_()
        self.desc_label.raise_()
        self.save_button.raise_()
        self.copytoclipboard_button.raise_()
        self.update_button.raise_()
        self.execute_button.raise_()
        self.module_combobox.raise_()
        self.rowcount_label.raise_()
        self.rowcount_lineedit.raise_()
        self.querydata_label.raise_()
        self.clearinputs_button.raise_()
        self.query_plaintextedit.raise_()
        self.stackedWidget.addWidget(self.searchentitypg)
        self.dbresultpg = QtGui.QWidget()
        self.dbresultpg.setObjectName(_fromUtf8("dbresultpg"))

        self.listTable = QtGui.QTableWidget(self.searchpg)
        self.listTable.setGeometry(QtCore.QRect(20, 66, 801, 371))
        self.listTable.setFrameShape(QtGui.QFrame.Box)
        self.listTable.setFrameShadow(QtGui.QFrame.Sunken)
        self.listTable.setMidLineWidth(0)
        #self.listTable.setEditTriggers(None)
        self.listTable.setProperty("showDropIndicator", True)
        self.listTable.setDragEnabled(False)
        self.listTable.setDefaultDropAction(QtCore.Qt.CopyAction)
        self.listTable.setAlternatingRowColors(False)
        self.listTable.setSelectionMode(QtGui.QAbstractItemView.ExtendedSelection)
        self.listTable.setSelectionBehavior(QtGui.QAbstractItemView.SelectRows)
        self.listTable.setTextElideMode(QtCore.Qt.ElideNone)
        self.listTable.setShowGrid(False)
        #self.listTable.setGridStyle(QtCore.Qt.SolidLine)
        self.listTable.setWordWrap(False)
        self.listTable.setCornerButtonEnabled(True)
        self.listTable.setObjectName(_fromUtf8("listTable"))
        self.listTable.verticalHeader().setVisible(False)
        self.listTable.resizeColumnsToContents()
       # self.listTable.horizontalHeader().setResizeMode(35,QtGui.QHeaderView.ResizeToContents)
        self.listTable.horizontalHeader().setCascadingSectionResizes(False)
        self.listTable.horizontalHeader().setStretchLastSection(False)
        self.listTable.verticalHeader().setDefaultSectionSize(20)
        self.listTable.setColumnCount(2)
        self.listTable.setColumnWidth(0,140)
        self.listTable.setColumnWidth(1,659)
##        self.listTable.setColumnWidth(2,59)
        #self.listTable.setSortingEnabled(True)
        
        sd= ['Module','Description','Time criticality']
        self.listTable.setHorizontalHeaderLabels(sd)
        
        
        #self.sqlTableWidget = QtGui.QTableWidget(self.dbresultpg)
        
        self.sqlTableWidget = QtGui.QTableView(self.dbresultpg)
        self.sqlTableWidget.setGeometry(QtCore.QRect(10, 20, 818, 421))
        self.sqlTableWidget.setFrameShape(QtGui.QFrame.NoFrame)
        self.sqlTableWidget.setFrameShadow(QtGui.QFrame.Sunken)
        self.sqlTableWidget.setMidLineWidth(0)
        #self.sqlTableWidget.setEditTriggers(QtGui.QAbstractItemView.AllEditTriggers)
        self.sqlTableWidget.setProperty("showDropIndicator", True)
        self.sqlTableWidget.setDragEnabled(False)
        self.sqlTableWidget.setDefaultDropAction(QtCore.Qt.CopyAction)
        
        self.sqlTableWidget.setAlternatingRowColors(True)
        self.sqlTableWidget.setSelectionMode(QtGui.QAbstractItemView.ExtendedSelection)
        self.sqlTableWidget.setTextElideMode(QtCore.Qt.ElideNone)
        self.sqlTableWidget.setGridStyle(QtCore.Qt.SolidLine)
        self.sqlTableWidget.setWordWrap(True)
        self.sqlTableWidget.setCornerButtonEnabled(True)
        self.sqlTableWidget.setObjectName(_fromUtf8("sqlTableWidget"))        
        self.sqlTableWidget.horizontalHeader().setCascadingSectionResizes(False)
        self.sqlTableWidget.horizontalHeader().setStretchLastSection(False)
##        self.sqlTableWidget.horizontalHeader().setResizeMode(QtGui.QHeaderView.Stretch)
        self.sqlTableWidget.verticalHeader().setDefaultSectionSize(20)
        self.model = QtGui.QStandardItemModel()
        self.sqlTableWidget.setModel(self.model)
        
        self.export_button = QtGui.QPushButton(self.dbresultpg)
        self.export_button.setGeometry(QtCore.QRect(737, 449, 91, 23))
        self.export_button.setObjectName(_fromUtf8("export_button"))
        
        self.term_button = QtGui.QPushButton(self.dbresultpg)
        self.term_button.setGeometry(QtCore.QRect(540, 450, 41, 18))
        self.term_button.setObjectName(_fromUtf8("term_button"))

        self.radio_csv = QtGui.QRadioButton('csv',self.dbresultpg)  ##radiobuttons
        self.radio_csv.setGeometry(QtCore.QRect(649, 450, 40, 18))
        self.radio_csv.setObjectName(_fromUtf8("radio_csv"))
        
        self.radio_xlsx = QtGui.QRadioButton('xlsx',self.dbresultpg)  ##radiobuttons
        self.radio_xlsx.setGeometry(QtCore.QRect(602, 450, 41, 18))
        self.radio_xlsx.setObjectName(_fromUtf8("radio_xlsx"))
        
        self.radio_txt = QtGui.QRadioButton('txt',self.dbresultpg)  ##radiobuttons
        self.radio_txt.setGeometry(QtCore.QRect(690, 450, 41, 18))
        self.radio_txt.setObjectName(_fromUtf8("radio_txt"))



        
        self.executeResult_label = QtGui.QLabel(self.dbresultpg)
        self.executeResult_label.setGeometry(QtCore.QRect(10, 0, 46, 13))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.executeResult_label.setFont(font)
        self.executeResult_label.setObjectName(_fromUtf8("executeResult_label"))
        self.runstatus03 = QtGui.QLabel(self.dbresultpg)
        self.runstatus03.setGeometry(QtCore.QRect(26, 450, 300, 16))
        self.runstatus03.setObjectName(_fromUtf8("runstatus03"))
        self.stackedWidget.addWidget(self.dbresultpg)
        self.prev_button = QtGui.QPushButton(self.centralwidget)
        self.prev_button.setGeometry(QtCore.QRect(759, 50, 41, 21))
        self.prev_button.setCheckable(False)
        self.prev_button.setObjectName(_fromUtf8("prev_button"))
        self.logoff_button = QtGui.QPushButton(self.centralwidget)
        self.logoff_button.setGeometry(QtCore.QRect(706, 50, 41, 21))
        self.logoff_button.setCheckable(False)
        self.logoff_button.setObjectName(_fromUtf8("logoff_button"))
        self.logoff_button.setText("Logout")
        self.logoff_button.setStyleSheet("color:#444444;")
        self.logoff_button.clicked.connect(self.logoff)        
        self.next_button = QtGui.QPushButton(self.centralwidget)
        self.next_button.setGeometry(QtCore.QRect(809, 50, 41, 21))
        self.next_button.setObjectName(_fromUtf8("next_button"))
        self.prev_button.setText("Back")
        self.next_button.setText("Next")
        self.scrollerwebview = QtWebKit.QWebView(self.centralwidget)
        self.scrollerwebview.setGeometry(QtCore.QRect(500, 20, 351, 15))
        self.scrollerwebview.setAutoFillBackground(True)
        self.scrollerwebview.setStyleSheet(_fromUtf8(""))        
        self.scrollerwebview.setObjectName(_fromUtf8("scrollerwebview"))
        self.scrollerwebview.setStyleSheet("background-color:rgb(117,176,176)")
        self.scrollerwebview.hide()
        self.mdcount_label = QtGui.QLabel(self.centralwidget)
        self.mdcount_label.setGeometry(QtCore.QRect(408, 21, 91, 16))
        self.mdcount_label.setObjectName(_fromUtf8("mdcount_label"))
        self.mdcount_label.setStyleSheet("color:rgb(57,57,173); font-weight:bold;")
        self.mdcount_label.setText("Module Counts:")
        
        
        self.stackedWidget.raise_()
        self.prev_button.raise_()
        self.next_button.raise_()
        self.scrollerwebview.raise_()
        self.login_group.raise_()
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtGui.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 871, 18))
        self.menubar.setLayoutDirection(QtCore.Qt.RightToLeft)
        self.menubar.setObjectName(_fromUtf8("menubar"))
        self.menuHelp = QtGui.QMenu(self.menubar)
        self.menuHelp.setObjectName(_fromUtf8("menuHelp"))
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtGui.QStatusBar(MainWindow)
        self.statusbar.setObjectName(_fromUtf8("statusbar"))
        MainWindow.setStatusBar(self.statusbar)
        self.statusbar.setStyleSheet("font-weight:bold")
        self.actionGuide = QtGui.QAction(MainWindow)
        self.actionGuide.setObjectName(_fromUtf8("actionGuide"))
        self.menuHelp.addAction(self.actionGuide)
        self.menubar.addAction(self.menuHelp.menuAction())
        
        self.progressBar = QtGui.QProgressBar(self.searchpg)
        
        self.progressBar.setGeometry(QtCore.QRect(360, 450, 118, 16))
        self.progressBar.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.progressBar.setAutoFillBackground(False)
        self.progressBar.setProperty("value", 0)
        self.progressBar.setAlignment(QtCore.Qt.AlignCenter)
        self.progressBar.setObjectName(_fromUtf8("progressBar"))
        
        self.progressBar2 = QtGui.QProgressBar(self.dbresultpg)
        
        self.progressBar2.setGeometry(QtCore.QRect(360, 450, 118, 16))
        self.progressBar2.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.progressBar2.setAutoFillBackground(False)
        self.progressBar2.setProperty("value", 0)
        self.progressBar2.setAlignment(QtCore.Qt.AlignCenter)
        self.progressBar2.setObjectName(_fromUtf8("progressBar2"))
        self.retranslateUi(MainWindow)
        self.stackedWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        MainWindow.setTabOrder(self.user_button, self.admin_button)
        MainWindow.setTabOrder(self.admin_button, self.adminpwd)
        MainWindow.setTabOrder(self.adminpwd, self.go_button)
        MainWindow.setTabOrder(self.go_button, self.searchkey_lineedit)
        MainWindow.setTabOrder(self.searchkey_lineedit, self.search_button)
        MainWindow.setTabOrder(self.search_button, self.add_button)
        MainWindow.setTabOrder(self.add_button, self.clear_button)
        MainWindow.setTabOrder(self.clear_button, self.resultspane_listwidget)
        MainWindow.setTabOrder(self.resultspane_listwidget, self.delete_button)
        MainWindow.setTabOrder(self.delete_button, self.module_combobox)
        MainWindow.setTabOrder(self.module_combobox, self.desc_lineedit)
        MainWindow.setTabOrder(self.desc_lineedit, self.query_plaintextedit)
        MainWindow.setTabOrder(self.query_plaintextedit, self.rowcount_lineedit)
        MainWindow.setTabOrder(self.rowcount_lineedit, self.execute_button)
        MainWindow.setTabOrder(self.execute_button, self.save_button)
        MainWindow.setTabOrder(self.save_button, self.copytoclipboard_button)
        MainWindow.setTabOrder(self.copytoclipboard_button, self.update_button)
        MainWindow.setTabOrder(self.update_button, self.clearinputs_button)
        MainWindow.setTabOrder(self.clearinputs_button, self.userId_lineedit)
        MainWindow.setTabOrder(self.userId_lineedit, self.password_lineedit)
        MainWindow.setTabOrder(self.password_lineedit, self.envdrop_combobox)
        MainWindow.setTabOrder(self.envdrop_combobox, self.sqlTableWidget)
        MainWindow.setTabOrder(self.sqlTableWidget, self.export_button)
        MainWindow.setTabOrder(self.export_button, self.prev_button)
        MainWindow.setTabOrder(self.prev_button, self.next_button)
        MainWindow.setTabOrder(self.next_button, self.scrollerwebview)
        
        self.audit_button = QtGui.QPushButton(self.searchentitypg)  ###########asasa###############
        self.audit_button.setGeometry(QtCore.QRect(450, 444, 75, 21))
        self.audit_button.setAutoDefault(False)
        self.audit_button.setObjectName(_fromUtf8("audit_button"))
        self.audit_button.setText("View audit")
        self.audit_button.clicked.connect(self.viewAudit)
        self.label_2.hide()
        self.pwdlabel.hide()
        self.adminpwd.hide()
        self.go_button.hide()
        self.add_button.hide()
        self.add_button.setEnabled(False)
        self.save_button.hide()
        self.update_button.hide()
        self.delete_button.hide()
        self.prev_button.clicked.connect(self.prevbutton)
        self.next_button.clicked.connect(self.nextbutton)
        self.search_button.clicked.connect(self.setSearchKey)
        self.listTable.doubleClicked.connect(self.printthis)
        self.resultspane_listwidget.doubleClicked.connect(self.printthis)
        self.admin_button.clicked.connect(self.adminbutton)
        self.execute_button.clicked.connect(self.diffProc) #execTri
        self.go_button.clicked.connect(self.gobutton)
        self.user_button.clicked.connect(self.usermode)
        self.delete_button.clicked.connect(self.delQuery)
        self.copytoclipboard_button.clicked.connect(self.copyClip)
        self.add_button.clicked.connect(self.changeInd)
        self.save_button.clicked.connect(self.addQu)
        self.update_button.clicked.connect(self.updQu)
        self.clearinputs_button.clicked.connect(self.lineReset)
        self.clear_button.clicked.connect(self.clearSearch)
        
##        self.prev_button.setStyleSheet("border:1px solid; border-radius:4px;")
##        self.next_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.search_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.resultspane_listwidget.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.admin_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.execute_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.go_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.user_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.delete_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.copytoclipboard_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.add_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.save_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.update_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.clearinputs_button.setStyleSheet("border:1px solid;border-radius:4px;")
##        self.clear_button.setStyleSheet("border:1px solid;border-radius:4px;")
        self.showStat()
        self.checkindex()

        
        
        
        self.export_button.clicked.connect(self.showMsgBox2)
        self.term_button.clicked.connect(self.terminate)
        
            
         
        
            
        k=len(MOD)
        for rows in range(k):
               self.module_combobox.addItem(_fromUtf8(""))
               self.module_combobox.setItemText(rows,MOD[rows])
        self.module_combobox.setMinimumHeight(10)
        self.module_combobox2.addItem(_fromUtf8(""))
        self.module_combobox2.setItemText(0,str("All"))
        for rows in range(k):
               self.module_combobox2.addItem(_fromUtf8(""))
               self.module_combobox2.setItemText(rows+1,MOD[rows])
        self.module_combobox2.setMinimumHeight(10)
        s=len(ENV)
        for rows in range(s):
               self.envdrop_combobox.addItem(_fromUtf8(""))
               self.envdrop_combobox.setItemText(rows,ENV[rows])
    def viewAudit(self):        
        import Audit
        self.audit=QtGui.QMainWindow()
        self.ui2=Audit.Ui_MainWindowAudit(queryDesc=self.desc_lineedit.text())
        self.ui2.setupUi(self.audit)
        self.audit.show()
          
    def diffProc(self): #difff
        self.stackedWidget.setCurrentIndex(3)
        self.model.clear()
        QtGui.qApp.processEvents()
        resultSet = []
        curdesc = ""
        errd =""
        qu= self.query_plaintextedit.toPlainText().replace(";","")
        qu =qu.rstrip()

        userid=self.userId_lineedit.text()
        pwd=self.password_lineedit.text()
        env=str(self.envdrop_combobox.currentText())
        numberOfRows=int(self.rowcount_lineedit.text())
##        q1 = multiprocessing.Manager().Queue()
##        q2 = multiprocessing.Manager().Queue()
##        q3 = multiprocessing.Manager().Queue()
####        q1 = queue.Queue()
####        q2 = queue.Queue()
####        q3 = queue.Queue()
##        q1.result =[]
##        q2z.curdesc=""
##        q1.err =""
        
##        p = multiprocessing.Process(target=self.workerobject.execQuery2, name="connectorThread",args=(self.userId_lineedit.text(),self.password_lineedit.text(),str(self.envdrop_combobox.currentText()),qu,int(self.rowcount_lineedit.text()),q1,q2,q3))
####        
####        p = threading.Thread(target=self.workerobject.execQuery2, name="connectorThread", args=(self.userId_lineedit.text(),self.password_lineedit.text(),str(self.envdrop_combobox.currentText()),qu,int(self.rowcount_lineedit.text()),q1,q2,q3))
####        p.daemon = True
        
####        p.start()
        
        print("proces started")
        
        self.workerobject =  WorkerThread(userid =userid, pwd=pwd ,env=env,queryStr=qu ,numberOfRows=numberOfRows)
        
        
        
        self.runstatus03.setText("")        
        self.workerobject.start()  
        self.statusbar.showMessage("Executing script...")
        
        
            
  
        
        
        try:
            QtGui.QMainWindow.connect(self.workerobject,QtCore.SIGNAL('fromWorker'),self.execQuery)
                  
        except BaseException as e:
            self.showMsgBox("Error",str(e))
        print("proces done")          
            
           
        
          
        
    def logoff(self):
        msgBox=QtGui.QMessageBox()
        msgBox.setIcon(QtGui.QMessageBox.Question)        
        msgBox.setWindowTitle("Confirm")
        msgBox.setText("Log Off the program?")
        msgBox.setStandardButtons(QtGui.QMessageBox.Yes |QtGui.QMessageBox.No)
        yesBt = msgBox.button(QtGui.QMessageBox.Yes)
        yesBt.setText('Yes')
        noBt = msgBox.button(QtGui.QMessageBox.No)
        noBt.setText('No')
        msgBox.exec_()
        if msgBox.clickedButton()==yesBt:
            self.clearSearch()
            self.lineReset()
            self.stackedWidget.setCurrentIndex(0)
            
            
            self.checkindex()
        
    def clearSearch(self):
        self.resultspane_listwidget.clear()        
        self.listTable.clear()        
        self.searchkey_lineedit.clear()
        self.runstatus01.clear()
        self.progressBar.setValue(0)
    def changeInd(self):
        
        self.stackedWidget.setCurrentIndex(2)
        self.module_combobox.setEnabled(True)
        self.lineReset()
        
    def copyClip(self):
        cb= QtGui.QApplication.clipboard()
        cb.clear(mode=cb.Clipboard)
        cb.setText(self.query_plaintextedit.toPlainText(),mode=cb.Clipboard)
        self.statusbar.showMessage("Query copied into clipboard")
    def checkindex(self): #checkindes
        
        if self.stackedWidget.currentIndex()==0:
            #and not(self.stackedWidget.currentIndex()==1 or 2 or 3):
            self.add_button.setEnabled(False)
            self.label.hide()
            self.scrollerwebview.hide()
            self.prev_button.hide()
            self.next_button.hide()                      
            self.mdcount_label.hide()
            self.logoff_button.hide()
        else:
            self.logoff_button.show()
            self.label.show()
            self.scrollerwebview.show()
            self.prev_button.show()
            self.next_button.show()
            self.scrollerwebview.show()            
            self.mdcount_label.show()

    def prevbutton(self):
        
        global conns
        conns =1
        if(self.stackedWidget.currentIndex()==1):
           
           self.stackedWidget.setCurrentIndex(3)

        else:
            self.stackedWidget.setCurrentIndex(self.stackedWidget.currentIndex() -1)
        self.checkindex()
    def nextbutton(self):
       
        if(self.stackedWidget.currentIndex()==3):
           self.stackedWidget.setCurrentIndex(1)
        else:
            self.stackedWidget.setCurrentIndex(self.stackedWidget.currentIndex() +1)
        #self.checkindex()
    def setSearchKey(self):
        
        global searchKey
        searchKey= self.searchkey_lineedit.text()
##        if searchKey ==' ' or searchKey=='' and self.combobox2.currentText() is not "All":
##    ##           self.SearchBar.setStyleSheet("QLineEdit {color:red;} QLineEdit:focus {color:black;} ")
##    ##           self.SearchBar.setText("Enter a Keyword")
##    ##           self.SearchBar.setStyleSheet("")
##            msgBox=QtGui.QMessageBox()
##            msgBox.setIcon(QtGui.QMessageBox.Information)            
##            msgBox.setWindowTitle("Message")
##            msgBox.setText("Please enter a keyword ")
##            msgBox.exec_()
##        else:
            
        if self.resultspane_listwidget.count() ==0:
            self.search()
        else:
            self.resultspane_listwidget.clear()
            self.search()

        
        searchKey=''
    def adminbutton(self):
        if self.add_button.isEnabled():
            self.label_2.setText("Already in admin mode")
            self.label_2.show()
            self.module_combobox.setGeometry(QtCore.QRect(80, 31, 250, 20))
            
        else:
            
            self.admin_button.hide()
            self.pwdlabel.show()
            self.adminpwd.show()
            self.adminpwd.setFocus()
            self.go_button.show()
        
    def gobutton(self):
        if self.adminpwd.text() == adminppwd:
            self.module_combobox.setGeometry(QtCore.QRect(80, 31, 250, 20))
            
            self.permi_label.show()
            self.permi_lineedit.show()
            self.label_2.hide()
            self.stackedWidget.setCurrentIndex(1)
            self.prev_button.show()
            self.next_button.show()
            self.audit_button.show()
            self.add_button.show()
            self.add_button.setEnabled(True)
            self.save_button.show()
            self.update_button.show()
            self.delete_button.show()
            self.admin_button.show()
            self.pwdlabel.hide()
            self.adminpwd.hide()
            self.go_button.hide()
            self.copytoclipboard_button.setGeometry(QtCore.QRect(110, 444, 101, 23))
            self.clearinputs_button.setGeometry(QtCore.QRect(340, 444, 75, 21))
            self.clear_button.setGeometry(QtCore.QRect(740, 20, 75, 21))
            self.checkindex()
            
        else:
            self.label_2.show()
            
            self.add_button.hide()
            self.save_button.hide()
            self.update_button.hide()
            self.delete_button.hide()
            
    def usermode(self):
        self.label_2.setText("Wrong password. Try again.")
        self.label_2.hide()
        self.adminpwd.clear()
        self.stackedWidget.setCurrentIndex(1)
        self.add_button.setEnabled(False)
        self.permi_label.hide()
        self.permi_lineedit.hide()
        self.add_button.hide()
        self.save_button.hide()
        self.audit_button.hide()
        self.update_button.hide()
        self.delete_button.hide()
        self.prev_button.show()
        self.next_button.show()
        self.module_combobox.setEnabled(False)
        self.copytoclipboard_button.setGeometry(QtCore.QRect(10, 444, 101, 23))
        self.clearinputs_button.setGeometry(QtCore.QRect(120, 444, 101, 23))
        self.clear_button.setGeometry(QtCore.QRect(641, 19, 90, 23))
        self.module_combobox.setGeometry(QtCore.QRect(80, 31, 501, 20))
        self.checkindex()
        
        
        

        
    def printthis(self):
        self.module_combobox.setEnabled(False)
        #self..hide()
        self.statusbar.showMessage("Getting item ...")
        #rr=self.listTable.selectedRows()
        d= self.listTable.currentRow()
        keyword2=self.listTable.item(d,1).text()
        keyword1=self.listTable.item(d,0).text()
        if keyword2!='':
##            sp = keyword2.split("  -  ")
            self.stackedWidget.setCurrentIndex(2)
            
            
            MDB =URL; DRV='{Microsoft Access Driver (*.mdb)}'
            con = pyodbc.connect('DRIVER={};DBQ={}'.format(DRV,MDB))

            cur = con.cursor()
##            keyword1 =sp[0]
##            keyword2=sp[1]
            #self.lineEdit_showModule.setText(keyword2)
            index =  self.module_combobox.findText(keyword1,QtCore.Qt.MatchFixedString)
            if index >=0:
                    self.module_combobox.setCurrentIndex(index)
            else:
                self.module_combobox.setCurrentIndex(self.module_combobox.findText("Others",QtCore.Qt.MatchFixedString))
                
            self.desc_lineedit.setText(keyword2)
            
            SQL="SELECT Module,Description,Query, Permission FROM Master where Master.[Module] ='"+keyword1+"' and Master.[Description] ='"+keyword2+"'"
            rows=cur.execute(SQL).fetchall()
            cur.close()
            con.close()
            self.statusbar.showMessage("")
            rowsList2 =rows
            userr=  getpass.getuser()#userr
            global origQuery
            origQuery = str(rowsList2[0][2])
            self.permi_lineedit.setText(str(rowsList2[0][3]))
            if str(rowsList2[0][3]) == 'All':
                self.query_plaintextedit.setPlainText(str(rowsList2[0][2]))
            elif userr in str(rowsList2[0][3]) :
                self.query_plaintextedit.setPlainText(str(rowsList2[0][2]))
            else:
                self.query_plaintextedit.setPlainText("")
                self.msgbox("x","Access denied","User '"+userr+"' does not have access to view this query.")
                
            
            
        
    def msgbox(self,typ,title,message):
        msgBox=QtGui.QMessageBox()
        if typ is "x":
            msgBox.setIcon(QtGui.QMessageBox.Critical) 
        elif typ is "i":
            msgBox.setIcon(QtGui.QMessageBox.Information) 
        else:
            msgBox.setIcon(QtGui.QMessageBox.Question) 
                   
        msgBox.setWindowTitle(title)
        msgBox.setText(message)
        msgBox.exec_()
        
    def addQu(self): 
       
        try:
            MDB =URL; DRV='{Microsoft Access Driver (*.mdb)}'
            con = pyodbc.connect('DRIVER={};DBQ={}'.format(DRV,MDB))
            cur = con.cursor()
            moduleStr=self.module_combobox.currentText()
            moduleStr=moduleStr.rstrip()
            descStr=self.desc_lineedit.text()
            descStr=descStr.rstrip()
            queryStr=self.query_plaintextedit.toPlainText()
            permiii=self.permi_lineedit.text()
            usrStr=getpass.getuser()
            if(moduleStr==''or descStr==''or queryStr=='' or len(moduleStr)<3 or len(descStr)<3 or len(queryStr)<3):
                    self.msgbox("x","Invalid Input","Please check the data in field(s)")
                    
            else:
                    
                    queryCom ="INSERT INTO Master ([Module],[Description],[Query],[UserId],[Permission]) VALUES(?,?,?,?,?)"
                    #+moduleStr+"','"+descStr+"','"+queryStr+"','"+usrStr+"','"+time.strftime("%m/%d/%Y")+"')"
                   
                    cur.execute(queryCom,moduleStr,descStr,queryStr,usrStr,permiii)#,time.strftime("%m/%d/%Y")
                    g=cur.rowcount
                    cur.commit()
                    con.commit()                                       
                    cur.close()
                    con.close()
                    self.lineReset()
                    if g !=0:
                        self.statusbar.showMessage("Saved!")
                        self.statusbar.setStyleSheet(_fromUtf8("color: green;font-weight: Bold;"))
                    else:
                        self.statusbar.showMessage("Save failed!")
                    
                    
        except BaseException as e:
            self.showMsgBox("Error",str(e))
    def showStat(self):           
        
        MDB = URL ; DRV='{Microsoft Access Driver (*.mdb)};Uid:jsaika01;pwd='
        con = pyodbc.connect('DRIVER={};DBQ={}'.format(DRV,MDB))

        cur = con.cursor()
        SQL="SELECT distinct [Module],count([Description]) FROM Master group by [Module]"
        rows=cur.execute(SQL).fetchall()
               
                   
        rowList=rows
##        item = QtGui.QListWidgetItem("Number of Queries loaded w.r.t Module: ")        
##        self.resultspane_listwidget.addItem(item)
        
        s1=""
        for i in range(0,len(rowList)):

            s1 =s1 +"<span style=color:white;>"+str( rowList[i][0]   )+":</span>&nbsp<span  style=color:white;);>"+str(rowList[i][1])+"</span>&nbsp&nbsp&nbsp&nbsp"
           
        
        SQL2="SELECT count(*) FROM Master "
        rows2=cur.execute(SQL2).fetchall()
        s2="<span style=color:white>TOTAL :"+str(rows2[0][0])+"</span>"      
        
        cur.close()
        con.close()
        self.scrollerwebview.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.scrollerwebview.setAttribute(QtCore.Qt.WA_OpaquePaintEvent,False)
        self.scrollerwebview.setHtml("<html  >"
                                     "<body style=margin:0 padding:0    >"
                                     "<font face=MS Shell Dlg 2 size=1  >" 
                                     "<strong>"
                                     "<marquee style=margin=0; padding=0;>"+s1+"   "+s2+"</marquee>"
                                     "</strong>"
                                     "</body>"
                                     "</html>")       
        #self.webView.scrollVerticalBar.hide()
        
            
        if len(rowList)==0:
            self.resultspane_listwidget.addItem(str("No data found"))
    def search(self):
##        dg= QtGui.QFileDialog()
##        dg.setFileMode(QtGui.QFileDialog.AnyFile)
##        name =""
##        dg.exec_()
##        name =dg.selectedFiles()
##        print (name)
            
        
        
        self.progressBar.setValue(0)
        self.resultspane_listwidget.clear()
        self.runstatus01.setText("")
        
        keyword11= searchKey
        
        
        MDB =URL; DRV='{Microsoft Access Driver (*.mdb)}'
        con = pyodbc.connect('DRIVER={};DBQ={}'.format(DRV,MDB))
        self.runstatus01.setText("Searching ...")
        cur = con.cursor()
        #searchentr
        keyword22=self.module_combobox2.currentText()
        if(keyword22=="All"):            
            SQL="SELECT Module,Description,Criticality FROM Master where lcase(Master.[Module]) like lcase('%"+keyword11+"%') or lcase(Master.[Description]) like lcase('%"+keyword11+"%') ORDER BY MODULE"
        else:
           # if (keyword11=="")
            SQL="SELECT Module,Description,Criticality FROM Master where lcase(Master.[Module]) like lcase('%"+keyword22+"%') and lcase(Master.[Description]) like lcase('%"+keyword11+"%') ORDER BY MODULE"
        rows=cur.execute(SQL).fetchall()
        self.resultspane_listwidget.hide()        
        font=QtGui.QFont()
        font.setBold(True)
        color =['Red','Orange','Yellow','Green','Blue']
        
        if len(rows) >0:
            self.runstatus01.setText("Fetching entries ...")
            
            self.listTable.horizontalHeader().setVisible(True)
            
            self.listTable.setColumnWidth(0,140)
            self.listTable.setColumnWidth(1,659)
##            self.listTable.setColumnWidth(2,59)
        
            sd= ['Module','Description','Time criticality']
            self.listTable.setHorizontalHeaderLabels(sd)
            self.listTable.setRowCount(len(rows))
            for i in  range(len(rows)):
                for j in range(len(cur.description)):
                    item = QtGui.QTableWidgetItem()
                    self.listTable.setItem(i, j, item)
                    
                    item.setText(str(rows[i][j]))
                    #row = row+str(rows[i][j])+";"
##                if rows[i][2]=='1':                    
##                    self.listTable.item(i,2).setBackground(QtGui.QColor('#4DADD1'))
##                    self.listTable.item(i,2).setForeground(QtGui.QColor('#4DADD1'))
##                elif rows[i][2]=='2':
##                    self.listTable.item(i,2).setBackground(QtGui.QColor('#7CBB52'))
##                    self.listTable.item(i,2).setForeground(QtGui.QColor('#7CBB52'))
##                elif rows[i][2]=='3':
##                    self.listTable.item(i,2).setBackground(QtGui.QColor('#D1D14D'))
##                    self.listTable.item(i,2).setForeground(QtGui.QColor('#D1D14D'))
##                elif rows[i][2]=='4':
##                    self.listTable.item(i,2).setBackground(QtGui.QColor('#CA864A'))
##                    self.listTable.item(i,2).setForeground(QtGui.QColor('#CA864A'))
##                elif rows[i][2]=='5':
##                    self.listTable.item(i,2).setBackground(QtGui.QColor('#B95D49'))
##                    self.listTable.item(i,2).setForeground(QtGui.QColor('#B95D49'))
##                else:
##                    self.listTable.item(i,2).setBackground(QtGui.QColor('#A1A1A1'))
##                    self.listTable.item(i,2).setForeground(QtGui.QColor('#A1A1A1'))
                self.listTable.item(i,0).setFont(font)                
                self.progressBar.setValue((i/len(rows))*100)
##                if ((int(len(rows)/5)) %5==0):
##                    QtGui.qApp.processEvents()
##            for k in range(len(rows)):
##                if rows[i][2]=='1':
##                    self.listTable.item(i,2).setBackground(QtGui.QtColor('Blue'))
##                elif rows[i][2]=='2':
##                    self.listTable.item(i,2).setBackground(QtGui.QtColor('Green'))
##                elif rows[i][2]=='3':
##                    self.listTable.item(i,2).setBackground(QtGui.QtColor('Yellow'))
##                elif rows[i][2]=='4':
##                    self.listTable.item(i,2).setBackground(QtGui.QtColor('Orange'))
##                elif rows[i][2]=='5':
##                    self.listTable.item(i,2).setBackground(QtGui.QtColor('Red'))
##                else:
##                    self.listTable.item(i,2).setBackground(QtGui.QtColor('Blue'))
            self.progressBar.setValue(100)
            self.runstatus01.setText("Showing "+str(len(rows))+" entries")
            #self.listTable.resizeColumnsToContents()
            
##            self.listTable.horizontalHeader().setResizeMode(1,QtGui.QHeaderView.ResizeToContents)                
        else:
            
            self.runstatus01.setText("")
            self.msgbox("i","No data found","Search keyword did not match with any entries.")
            self.progressBar.setValue(0)
			
        cur.close()
        con.close()
        
    def exportToTxt(self):
        try:
            if  self.model.columnCount()!=0 or self.model.rowCount()!=0:
                self.runstatus03.setText("Exporting, please wait...")
                tick1=time.clock()

                desfile ='Export_'+str(time.strftime('%y-%m-%d_%H%M%S'))+'.txt'
                outputfile = open(desfile,'w')
                output = csv.writer(outputfile,delimiter=' ')
                coll=[]
                for a in range(self.model.columnCount()):
                    coll.append(self.model.horizontalHeaderItem(a).text())
                output.writerow(coll)
                QtGui.qApp.processEvents()
                self.progressBar.setValue(0)
                QtGui.qApp.processEvents()

                roww=[[self.model.item(i,j).text() for j in range(self.model.columnCount()) ] for i in range(self.model.rowCount())]
                
##                for i in range(self.model.rowCount()):
##                    for j in range(self.model.columnCount()):
                      
##                       if self.model.item(i,j) is None:
##                           roww.append("")
##                           
##                       else:
##                           roww.append(self.model.item(i,j).text())
                    
                for index in roww:
                    self.progressBar.setValue((i/self.model.rowCount())*100)
                   # print (index)
                    output.writerow(index)
                    
                self.progressBar.setValue(100)
                QtGui.qApp.processEvents()

               
                outputfile.close()
                tick2=time.clock()
                self.progressBar.setValue(100)
                self.runstatus03.setText("Done! Data saved in "+os.getcwd()+"\\"+desfile)
                self.statusbar.showMessage("Export completed in "+str(round((tick2-tick1),3))+" seconds")
                self.runstatus03.setStyleSheet(_fromUtf8("color: Green;"))
            else:
                self.runstatus03.setText("Data not saved: No data found in the table.")
                self.runstatus03.setStyleSheet(_fromUtf8("color:Red;"))
                
            
        except BaseException as e:            
            self.runstatus03.setText("Error: "+str(e))
            self.runstatus03.setStyleSheet(_fromUtf8("color: Red;"))
       
        
        
    def exportToCsv(self):    
        try:
            if  self.model.columnCount()!=0 or self.model.rowCount()!=0:
                self.runstatus03.setText("Exporting, please wait...")
                tick1=time.clock()

                desfile ='Export_'+str(time.strftime('%y-%m-%d_%H%M%S'))+'.csv'
                outputfile = open(desfile,'w')
                output = csv.writer(outputfile,dialect='excel',quotechar="'",quoting=csv.QUOTE_NONNUMERIC)
                coll=[]
                for a in range(self.model.columnCount()):
                    coll.append(self.model.horizontalHeaderItem(a).text())
                output.writerow(coll)
                QtGui.qApp.processEvents()
                self.progressBar.setValue(0)
                QtGui.qApp.processEvents()

                roww=[[self.model.item(i,j).text() for j in range(self.model.columnCount()) ] for i in range(self.model.rowCount())]
                
##                for i in range(self.model.rowCount()):
##                    for j in range(self.model.columnCount()):
                      
##                       if self.model.item(i,j) is None:
##                           roww.append("")
##                           
##                       else:
##                           roww.append(self.model.item(i,j).text())
                    
                for index in roww:
                    self.progressBar.setValue((i/self.model.rowCount())*100)
                   # print (index)
                    output.writerow(index)
                    
                self.progressBar.setValue(100)
                QtGui.qApp.processEvents()

               
                outputfile.close()
                tick2=time.clock()
                self.progressBar.setValue(100)
                self.runstatus03.setText("Done! Data saved in "+os.getcwd()+"\\"+desfile)
                self.statusbar.showMessage("Export completed in "+str(round((tick2-tick1),3))+" seconds")
                self.runstatus03.setStyleSheet(_fromUtf8("color: Green;"))
            else:
                self.runstatus03.setText("Data not saved: No data found in the table.")
                self.runstatus03.setStyleSheet(_fromUtf8("color:Red;"))
                
            
        except BaseException as e:            
            self.runstatus03.setText("Error: "+str(e))
            self.runstatus03.setStyleSheet(_fromUtf8("color: Red;"))
       
        
        

        
    def exportToExcel(self):
        try:
            if  self.model.columnCount()!=0 or self.model.rowCount()!=0:
                self.runstatus03.setText("Exporting, please wait...")
                tick1=time.clock()
                wb=  Workbook()
                desfile ='Export_'+str(time.strftime('%y-%m-%d_%H%M%S'))+'.xlsx'
                ws1 = wb.active
                ws1.title='Export'
                
                for a in range(self.model.columnCount()):
                    ws1.cell(row=1, column=a+1).value=self.model.horizontalHeaderItem(a).text()
                QtGui.qApp.processEvents()
                self.progressBar.setValue(0)
                QtGui.qApp.processEvents()
                for i in range(self.model.rowCount()):
                    for j in range(self.model.columnCount()):
                       # print (self.model.item(i,j).text())
                       if self.model.item(i,j) is None:
                           ws1.cell(row=i+2, column=j+1).value=""
                           
                       else:
                           ws1.cell(row=i+2, column=j+1).value=self.model.item(i,j).text()
                    self.progressBar.setValue((i/self.model.rowCount())*100)
                self.progressBar.setValue(100)
                QtGui.qApp.processEvents()

                ws2= wb.create_sheet()
                ws2.title = 'SQL'
                ws2.cell(row=1, column=1).value=self.query_plaintextedit.toPlainText()

            
                wb.save(filename = desfile)
                tick2=time.clock()
                self.progressBar.setValue(100)
                self.runstatus03.setText("Done! Data saved in "+os.getcwd()+"\\"+desfile)
                self.statusbar.showMessage("Export completed in "+str(round((tick2-tick1),3))+" seconds")
                self.runstatus03.setStyleSheet(_fromUtf8("color: Green;"))
            else:
                self.runstatus03.setText("Data not saved: No data found in the table.")
                self.runstatus03.setStyleSheet(_fromUtf8("color:Red;"))
                
            
        except BaseException as e:            
            self.runstatus03.setText("Error: "+str(e))
            self.runstatus03.setStyleSheet(_fromUtf8("color: Red;"))
         
    def delQuery(self):   
        
            sel =self.listTable.selectionModel().selectedRows()
            if len(sel)>0:
                self.confirmDelete()                
            else:
                self.runstatus01.setText("Please select items first.") 
       
                       
          
    def confirmDelete(self): #qqq
        self.listTable.setSortingEnabled(False)
        sel =self.listTable.selectionModel().selectedRows()
        print (sel)
        msgBox=QtGui.QMessageBox()
        msgBox.setIcon(QtGui.QMessageBox.Question)
        #msgBox.setAttribute(QtCore.Qt.WA_DeleteOnClose)
        #msgBox.setWindowModality(QtCore.Qt.NonModal)
        msgBox.setWindowTitle("Confirm")
        msgBox.setText("Are you sure you want to delete? This operation cannot be undone.")
        msgBox.setStandardButtons(QtGui.QMessageBox.Yes |QtGui.QMessageBox.No)
        yesBt = msgBox.button(QtGui.QMessageBox.Yes)
        yesBt.setText('Yes')
        noBt = msgBox.button(QtGui.QMessageBox.No)
        noBt.setText('No')
        msgBox.exec_()
        if msgBox.clickedButton()==yesBt:               
            
                #self.runstatus03.setStyleSheet(_fromUtf8("color: Blue;"))
                MDB =URL; DRV='{Microsoft Access Driver (*.mdb)}'
                con = pyodbc.connect('DRIVER={};DBQ={}'.format(DRV,MDB))
                cur = con.cursor()
                n = 0
                for i in range(len(sel)):                
                    queryCom ="Delete from Master where Description=?"            
                    cur.execute(queryCom,self.listTable.item(sel[i].row(),1).text())
                    n =n+cur.rowcount
                    self.listTable.removeRow(sel[i].row())
                    #print (sel[i].data().split("  -  ")[0])                
                cur.commit()
                con.commit()
                
                cur.close()
                con.close()
                
                QtGui.qApp.processEvents()                    
                self.statusbar.showMessage("Deleted "+str(n)+" entries")
                n=0
        else:
            self.statusbar.showMessage("Operation cancelled.")
            
    
             
             
                    
        
         
    def execQuery(self,resultSet,curdesc,err,timdiff): ##oraexe
        
        if self.userId_lineedit.text() or self.password_lineedit.text() is not "" :
            if(err is not "" ):
                self.showMsgBox("Error",err)
            if (len(resultSet)==0):
                self.runstatus03.setText("No data returned")
            
            self.stackedWidget.setCurrentIndex(3)
            QtGui.qApp.processEvents() 
            
            
            
            try:                
                self.model.clear()
                self.statusbar.showMessage("")
                QtGui.qApp.processEvents()                
               # conn =cx_Oracle.connect(usern,pwd,constr)
                
                
                QtGui.qApp.processEvents()  
                #curr = conn.cursor()
                #qrystr = self.query_plaintextedit.toPlainText().replace(";","")
                #qrystr =qrystr.rstrip()
                
##                while True:
##                    time.sleep(1)
##                    QtCore.QCoreApplication.processEvents()
##                    if( r==1):
##                        break
##                curr.execute(qrystr)                       
##                numberOfRows= int(self.rowcount_lineedit.text()) 
                             
                 
##                col_names = [] 
                
                self.progressBar2.minimum = 1
                self.progressBar2.maximum = len(resultSet)
                
                
                curdesc= curdesc[:-1]    
                self.model.setHorizontalHeaderLabels(curdesc.split(";"))
                colss= curdesc.split(";")
                
                numCol=len(colss)
                
                mrow =QtGui.QStandardItem()
                if (len(resultSet)>0) :
                        self.runstatus03.setText("Fetching results...")
                        g= []
                        for i in  range(len(resultSet)):               
                            for j in range(len(colss)):
                              rrow =QtGui.QStandardItem(str(resultSet[i][j]))
                              g.append(rrow)                    
                            self.model.appendRow(g)
                            g=[]
    ##                        if (i % (len(resultSet)/5)==0):
    ##                            QtCore.QCoreApplication.processEvents()
                            
                            self.progressBar2.setValue((i/len(resultSet))*100)
                        QtCore.QCoreApplication.processEvents()                
                        self.progressBar2.setValue(100)
                        QtGui.qApp.processEvents()                                                  
                        self.runstatus03.setText("Task completed in "+timdiff+" seconds")                    
                        self.statusbar.showMessage("Displaying "+str(len(resultSet))+" rows"+", "+str(len(colss))+" columns")
                        
                else:
                    self.runstatus03.setText("")
                    self.statusbar.showMessage("No data returned")
                
            except BaseException as e:
                    self.showMsgBox("Error",str(e))
        else:
            self.msgbox("x","Message","Username or password is blank")
            
    def showMsgBox(self,title,message):
        
        msgBox=QtGui.QMessageBox()
        msgBox.setIcon(QtGui.QMessageBox.Critical)
        
        msgBox.setWindowTitle(title)
        msgBox.setText(message)
        msgBox.exec_()
        
    def updQu(self):
        
        MDB =URL; DRV='{Microsoft Access Driver (*.mdb)}'
        con = pyodbc.connect('DRIVER={};DBQ={}'.format(DRV,MDB))
        cur = con.cursor()
        moduleStr=self.module_combobox.currentText()
        descStr=self.desc_lineedit.text()
        queryStr=self.query_plaintextedit.toPlainText() #USe this for query input oracle
        permii= self.permi_lineedit.text()
        usrStr=getpass.getuser()
        
        if(moduleStr==''or descStr==''or queryStr=='' or len(moduleStr)<3 or len(descStr)<3 or len(queryStr)<3):
                self.msgbox("x","Invalid inputs","Please check the data in field(s)")
                
        else:                #ueryStr.replace("'","''")
                queryCom ="UPDATE Master set [Query]=?,[UserId]=?,[Permission]=? where Module=? and Description=?"
                #+moduleStr+"','"+descStr+"','"+queryStr+"','"+usrStr+"','"+time.strftime("%m/%d/%Y")+"')"               
                cur.execute(queryCom,queryStr,usrStr,permii,moduleStr,descStr)
                f=cur.rowcount
                cur.commit()
                con.commit()
                checkqry= "SELECT count(QueryDesc) FROM Audit where QueryDesc=?"
                checkNum = cur.execute(checkqry,descStr).fetchall()
                print (checkNum)
                atype= "Update"
                if int(checkNum[0][0]) <= 4:
                    queryCom2 ="INSERT INTO Audit ([Module],[QueryDesc],[UserId],[BeforeVal],[AfterVal],[Type]) VALUES(?,?,?,?,?,?)"
                    cur.execute(queryCom2,moduleStr,descStr,usrStr,origQuery,queryStr,atype)
                    cur.commit()
                    con.commit()
                
                
                cur.close()
                con.close()
                self.lineReset()
##                time.sleep(0.5)
                if f!=0:
                    self.statusbar.showMessage("Updated!")
                    self.statusbar.setStyleSheet("color:green")
                else:
                    self.statusbar.showMessage("Update failed!")
                    self.statusbar.setStyleSheet("color:orange")
                    
##                
    def lineReset(self):
        self.query_plaintextedit.clear()
        self.desc_lineedit.clear()
        self.module_combobox.setCurrentIndex(0)
    def terminate(self): #terter
        self.workerobject.terminate()
        if not self.workerobject.isRunning():
            self.statusbar.showMessage("Query execution cancelled")
        else:
            self.statusbar.showMessage("Still executing...")        
            
        
        
    def showMsgBox2(self): #export_button trigcx
        self.runstatus03.setStyleSheet(_fromUtf8("color:black;"))
        if  self.model.columnCount()!=0 or self.model.rowCount()!=0:
            if not (self.radio_xlsx.isChecked() or self.radio_csv.isChecked() or self.radio_txt.isChecked() ):
                self.msgbox("x","Message","Please select the file format")
            else:
            
                msgBox=QtGui.QMessageBox()
                msgBox.setIcon(QtGui.QMessageBox.Question)
                #msgBox.setAttribute(QtCore.Qt.WA_DeleteOnClose)
                #msgBox.setWindowModality(QtCore.Qt.NonModal)
                msgBox.setWindowTitle("Confirm")
                msgBox.setText("Are you sure you want to export the data?")
                msgBox.setStandardButtons(QtGui.QMessageBox.Yes |QtGui.QMessageBox.No)            
                yesBt = msgBox.button(QtGui.QMessageBox.Yes)
                yesBt.setText('Yes')
                noBt = msgBox.button(QtGui.QMessageBox.No)
                noBt.setText('No')
                msgBox.exec_()
                if msgBox.clickedButton()==yesBt:
                    if(self.radio_xlsx.isChecked()):                    
                        self.exportToExcel()
                    elif(self.radio_csv.isChecked()):
                        self.exportToCsv()
                    elif(self.radio_txt.isChecked()):
                        self.exportToTxt()
                
                
        else:
            self.runstatus03.setText("Export failed: No data found in the table.")
            self.runstatus03.setStyleSheet(_fromUtf8("color:Red;"))
        

            
    def retranslateUi(self, MainWindow):
        MainWindow.setWindowTitle(_translate("MainWindow", "Test Data Mining tool", None))
        
        self.login_group.setTitle(_translate("MainWindow", "Login", None))
        self.user_button.setText(_translate("MainWindow", "User", None))
        self.admin_button.setText(_translate("MainWindow", "Admin", None))
        self.pwdlabel.setText(_translate("MainWindow", "Password:", None))
        self.go_button.setText(_translate("MainWindow", "Go", None))
        self.label_2.setText(_translate("MainWindow", "Wrong password. Try again", None))
        self.clear_button.setText(_translate("MainWindow", "Clear", None))
        self.search_button.setText(_translate("MainWindow", "Search", None))
        self.add_button.setText(_translate("MainWindow", "Add", None))
        #self.searchkey_lineedit.setText(_translate("MainWindow", "Sample search key", None))
        self.resultspane_label.setText(_translate("MainWindow", "Results:", None))
        self.delete_button.setText(_translate("MainWindow", "Delete", None))
        self.searchhere_label.setText(_translate("MainWindow", "Start:", None))
        #self.runstatus01.setText(_translate("MainWindow", "Showing", None))
        self.query_label.setText(_translate("MainWindow", "Query/Data:", None))
        self.module_label.setText(_translate("MainWindow", "Module:", None))
        #self.desc_lineedit.setText(_translate("MainWindow", "desc_text", None))
        self.desc_label.setText(_translate("MainWindow", "Description:", None))
        self.dbcreds_group.setTitle(_translate("MainWindow", "DB Credentials", None))
        self.pwd_label.setText(_translate("MainWindow", "Password", None))
  
        self.usrname_label.setText(_translate("MainWindow", "Username", None))
        self.env_label.setText(_translate("MainWindow", "Environment", None))
        self.save_button.setText(_translate("MainWindow", "Save", None))
        self.copytoclipboard_button.setText(_translate("MainWindow", "Copy to Clipboard", None))
        self.update_button.setText(_translate("MainWindow", "Update", None))
        self.execute_button.setText(_translate("MainWindow", "Execute", None))
        self.rowcount_label.setText(_translate("MainWindow", "Row count:", None))
        self.rowcount_lineedit.setText(_translate("MainWindow", "100", None))
        self.querydata_label.setText(_translate("MainWindow", "Query/ Data:", None))
        #self.mdcount_label.setText(_translate("MainWindow", "Module Counts:", None))
        self.clearinputs_button.setText(_translate("MainWindow", "Clear", None))        
        self.export_button.setText(_translate("MainWindow", "Export", None))
        self.term_button.setText(_translate("MainWindow", "Stop", None))
        self.executeResult_label.setText(_translate("MainWindow", "Results:", None))
        self.runstatus03.setText(_translate("MainWindow", "Ready", None))
        #self.prev_button.setText(_translate("MainWindow", "Prev", None))
        #self.next_button.setText(_translate("MainWindow", "Next", None))
        #self.menuHelp.setTitle(_translate("MainWindow", "Help", None))
        #self.actionGuide.setText(_translate("MainWindow", "Guide", None))


class qCustom (QtGui.QMainWindow):
    def closeEvent(self,event):
        msg= "Are you sure you want to exit the program?"
        reply=QtGui.QMessageBox.question(self,'Message',msg, QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
        if reply == QtGui.QMessageBox.Yes:
            
            event.accept()
            
        else:
            event.ignore()

class WorkerThread(QtCore.QThread): #workercls

    def __init__(self,userid,pwd,env,queryStr,numberOfRows, parent=None):
        QtCore.QThread.__init__(self,parent)
##        self.signalvar = QtCore.pyqtSignal()
        self.userid =userid
        self.pwd=pwd
        self.env=env
        self.queryStr=queryStr
        self.numberOfRows=numberOfRows

    def run(self):
        
        conn =None
        curr = None
        resultSet=[]
        curdesc=""
        err=""
        tick1=time.clock()
        try:
            
            conn =cx_Oracle.connect(self.userid,self.pwd,self.env)
            curr = conn.cursor()
            qrystr =self.queryStr.rstrip()    
            curr.execute(qrystr)
            result= curr.fetchmany(numRows=self.numberOfRows)
            
            for i in range(len(curr.description)):
                curdesc= curdesc+str(curr.description[i][0])+";"               
            
            resultSet = result
            

        except BaseException as e:                 
             err =str(e)
        finally:            
            if conn and curr is not None:
                curr.close()
                conn.close()
            tick2=time.clock()
            timdiff=str(round((tick2-tick1),3))
            self.emit(QtCore.SIGNAL('fromWorker'),resultSet,curdesc,err,timdiff)
            return

    
        
                
if __name__ == "__main__":
   
    app = QtGui.QApplication(sys.argv)
    MainWindow = qCustom()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()

   # multiprocessing.freeze_support()
    sys.exit(app.exec_())



